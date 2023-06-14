import os
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from decimal import Decimal, InvalidOperation, ROUND_DOWN
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
import pandas as pd


def per_to_f(s):
    return float(s.rstrip('%')) / 100


def s_decimal(string):
    try:
        return Decimal(string)
    except InvalidOperation:
        return None


class ExcelWriter:
    def write_datas(self, sporttery_datas, okooo_datas):
        cur_date = datetime.now()
        date_str = cur_date.strftime('%Y-%m-%d')
        for sporttery_data in sporttery_datas:
            for okooo_data in okooo_datas:
                if (sporttery_data['match_num'] == okooo_data['match_num']) and \
                        (sporttery_data['business_date'] == date_str):
                    self.write_data(sporttery_data, okooo_data)


    def gen_line_chart(self, ws, sporttery_data, image_path):
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False

        start_row = 3
        end_row = ws.max_row
        step = 3

        # get time datas
        time_datas = [ws['A'][i].value for i in range(start_row - 1, end_row, step)]
        time_objs = pd.to_datetime(time_datas, format='%H:%M')
        time_objs = time_objs.to_list()

        for i in range(1, len(time_objs)):
            if time_objs[i] < time_objs[i - 1]:
                time_objs[i] += timedelta(days=1)

        time_objs = pd.DatetimeIndex(time_objs)

        # get payout datas
        payout_h_datas = [float(ws['AA'][i].value) for i in range(start_row - 1, end_row, step)]
        payout_d_datas = [float(ws['AB'][i].value) for i in range(start_row - 1, end_row, step)]
        payout_a_datas = [float(ws['AC'][i].value) for i in range(start_row - 1, end_row, step)]

        # draw
        fig, ax = plt.subplots()
        ax.plot(time_objs, payout_h_datas, color='blue', marker='o', markersize=4, label='胜')
        ax.plot(time_objs, payout_d_datas, color='#FFA500', marker='o', markersize=4, label='平')
        ax.plot(time_objs, payout_a_datas, color='gray', marker='o', markersize=4, label='负')

        date_format = DateFormatter('%H:%M')
        ax.xaxis.set_major_formatter(date_format)
        fig.autofmt_xdate()
        plt.ylim(bottom=0, top=max(max(payout_h_datas), max(payout_a_datas), max(payout_a_datas)) + 0.5)

        title1 = sporttery_data['match_num'] + ' ' + sporttery_data['home_team'] + 'vs' + sporttery_data['away_team']
        title2 = '胜平负赔付率'
        plt.title(f'{title1}\n{title2}')
        plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=3)

        plt.savefig(image_path)
        plt.close()

    def write_data(self, sporttery_data, okooo_data):
        excel_filename = sporttery_data['business_date'] + '.xlsx'
        excel_dir = '../excels/' + excel_filename[0:7] + '/'
        excel_path = os.path.join(excel_dir, excel_filename)

        os.makedirs(excel_dir, exist_ok=True)

        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()

        sheet_name = sporttery_data['match_num'] + sporttery_data['home_team'] \
                     + 'vs' + sporttery_data['away_team']

        if sheet_name in wb.sheetnames:
            ws = wb.get_sheet_by_name(sheet_name)
        else:
            ws = wb.create_sheet(sheet_name)
            if 'Sheet' in wb.sheetnames:
                default_sheet = wb['Sheet']
                wb.remove(default_sheet)

            ws.column_dimensions['B'].width = 2

            ws.merge_cells('C1:E1')
            ws['C1'] = '赔率'

            ws.column_dimensions['F'].width = 2

            ws.merge_cells('G1:I1')
            ws['G1'] = '官网支持率'

            ws.column_dimensions['J'].width = 2

            ws.merge_cells('K1:M1')
            ws['K1'] = '必发'

            ws.column_dimensions['N'].width = 2

            ws.merge_cells('O1:Q1')
            ws['O1'] = '澳客保存率'

            ws.column_dimensions['R'].width = 2

            ws.merge_cells('S1:U1')
            ws['S1'] = '澳客人气'

            ws.column_dimensions['V'].width = 2

            ws.merge_cells('W1:Y1')
            ws['W1'] = '平均支持率'

            ws.column_dimensions['Z'].width = 2

            ws.merge_cells('AA1:AC1')
            ws['AA1'] = '赔付'

            ws.column_dimensions['AD'].width = 2

            # set head line 2
            ws['A2'] = sporttery_data['match_num']

            ws['C2'] = '主胜'
            ws['D2'] = '平'
            ws['E2'] = '客胜'

            ws['G2'] = '主胜'
            ws['H2'] = '平'
            ws['I2'] = '客胜'

            ws['K2'] = '主胜'
            ws['L2'] = '平'
            ws['M2'] = '客胜'

            ws['O2'] = '主胜'
            ws['P2'] = '平'
            ws['Q2'] = '客胜'

            ws['S2'] = '主胜'
            ws['T2'] = '平'
            ws['U2'] = '客胜'

            ws['W2'] = '主胜'
            ws['X2'] = '平'
            ws['Y2'] = '客胜'

            ws['AA2'] = '主胜'
            ws['AB2'] = '平'
            ws['AC2'] = '客胜'

        print(ws.max_row)

        cur_time = datetime.now().strftime('%H:%M')

        if okooo_data['betfair_index_h'] == '':
            #return
            # for test
            okooo_data['betfair_index_h'] = '0.3333'
            okooo_data['betfair_index_d'] = '0.3333'
            okooo_data['betfair_index_a'] = '0.3333'

        average_sr_h = s_decimal(sporttery_data['had_hsr']) * Decimal('0.35') + \
            s_decimal(okooo_data['betfair_index_h']) * Decimal('0.1') + \
            s_decimal(okooo_data['save_rate_h']) * Decimal('0.35') + \
            s_decimal(okooo_data['popularity_rate_h']) * Decimal('0.2')
        average_sr_h = average_sr_h.quantize(Decimal('0.0000'), rounding=ROUND_DOWN)

        average_sr_d = s_decimal(sporttery_data['had_dsr']) * Decimal('0.35') + \
            s_decimal(okooo_data['betfair_index_d']) * Decimal('0.1') + \
            s_decimal(okooo_data['save_rate_d']) * Decimal('0.35') + \
            s_decimal(okooo_data['popularity_rate_d']) * Decimal('0.2')
        average_sr_d = average_sr_d.quantize(Decimal('0.0000'), rounding=ROUND_DOWN)

        average_sr_a = s_decimal(sporttery_data['had_asr']) * Decimal('0.35') + \
            s_decimal(okooo_data['betfair_index_a']) * Decimal('0.1') + \
            s_decimal(okooo_data['save_rate_a']) * Decimal('0.35') + \
            s_decimal(okooo_data['popularity_rate_a']) * Decimal('0.2')
        average_sr_a = average_sr_a.quantize(Decimal('0.0000'), rounding=ROUND_DOWN)

        payout_r_h = average_sr_h * Decimal(sporttery_data['had_h'])
        payout_r_h = payout_r_h.quantize(Decimal('0.0000'), rounding=ROUND_DOWN)
        payout_r_d = average_sr_d * Decimal(sporttery_data['had_d'])
        payout_r_d = payout_r_d.quantize(Decimal('0.0000'), rounding=ROUND_DOWN)
        payout_r_a = average_sr_a * Decimal(sporttery_data['had_a'])
        payout_r_a = payout_r_a.quantize(Decimal('0.0000'), rounding=ROUND_DOWN)

        data1 = [cur_time, None,
                 float(sporttery_data['had_h']), float(sporttery_data['had_d']), float(sporttery_data['had_a']), None,
                 per_to_f(sporttery_data['had_hsr']), per_to_f(sporttery_data['had_dsr']),
                 per_to_f(sporttery_data['had_asr']), None,
                 s_decimal(okooo_data['betfair_index_h']), s_decimal(okooo_data['betfair_index_d']),
                 s_decimal(okooo_data['betfair_index_a']), None,
                 s_decimal(okooo_data['save_rate_h']), s_decimal(okooo_data['save_rate_d']),
                 s_decimal(okooo_data['save_rate_a']), None,
                 s_decimal(okooo_data['popularity_rate_h']), s_decimal(okooo_data['popularity_rate_d']),
                 s_decimal(okooo_data['popularity_rate_a']), None,
                 average_sr_h, average_sr_d, average_sr_a, None,
                 payout_r_h, payout_r_d, payout_r_a, None]
        data2 = [None, None,
                 float(sporttery_data['hhad_h']), float(sporttery_data['hhad_d']), float(sporttery_data['hhad_a']),
                 None,
                 per_to_f(sporttery_data['hhad_hsr']), per_to_f(sporttery_data['hhad_dsr']),
                 per_to_f(sporttery_data['hhad_asr']), None]

        if (ws.max_row > 2):
            ws.append([])
        ws.append(data1)
        ws.append(data2)

        num_rows = ws.max_row
        ws.merge_cells(start_row=num_rows - 1, start_column=1, end_row=num_rows, end_column=1)

        image_dir = '../images/' + sporttery_data['business_date'] + '/'
        os.makedirs(image_dir, exist_ok=True)
        image_name = sheet_name + '.png'
        image_path = os.path.join(image_dir, image_name)
        self.gen_line_chart(ws, sporttery_data, image_path)

        for img in ws._images:
            print("remove a img")
            ws._images.remove(img)

        img = Image(image_path)
        ws.add_image(img, 'AE2')

        alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        wb.save(excel_path)
